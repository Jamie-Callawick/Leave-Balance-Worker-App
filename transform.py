"""
Smartly → Simplifi Leave Balance Transformer
=============================================
Reads a Smartly leave balance export CSV and produces a Simplifi-formatted
Excel file ready for Staff Import.

Included leave types:
  - Annual Leave  → balance taken from "Balance (Hours)"
  - Sick          → balance taken from "Balance" (days)

Usage:
  python transform.py <input_csv> [output_xlsx]

  If output_xlsx is omitted, the file is saved alongside the input CSV
  with "_simplifi_import" appended to the name.

Example:
  python transform.py "Leave Balances PPE 12_04_26.csv"
  python transform.py "Leave Balances PPE 12_04_26.csv" "simplifi_import.xlsx"
"""

import sys
import csv
import io
import os
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Install it with:  pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

# Maps Smartly leave type name → (simplifi_type_code, balance_column)
# balance_column is either "Balance (Hours)" or "Balance"
LEAVE_TYPE_MAP = {
    "Annual Leave": ("Annual Leave", "Balance (Hours)"),
    "Sick":         ("Sick",         "Balance"),
}

# Number of metadata rows to skip at the top of the Smartly CSV
# (Pay Groups line, Leave Types line, Departments line, Period End Date line, blank line)
SMARTLY_HEADER_SKIP = 5

# Simplifi template column names (must match the template exactly)
SIMPLIFI_COLUMNS = [
    "First Name", "Last Name", "Phone", "Email Address",
    "Email Address (Alternative)", "Tax Number", "Address (Complete)",
    "Street Number", "Street Name", "Suburb", "Town/City", "Region/State",
    "Post Code", "Country", "Is Being Manually Maintained",
    "Is Included in Payroll", "Is Rostered  Employee", "Is Active Employee",
    "Employee Cost Code", "Employee Code", "Job Role(s)", "Department",
    "Home Department", "Head Office Code", "Pay Roll - Pay Roll Code",
    "Pay Roll - External Timesheet Provider Code",
    "Pay Roll - Last Date Employment", "Pay Roll - Calendar",
    "Pay Roll - Job Role for Pay Rate", "Pay Roll - Pay Rate",
    "Pay Roll - Pay Rate Valid From Date", "Leave - Balance Owing",
    "Leave - Type Code", "Contracted Hours - From Date",
    "Contracted Hours - Effective To Date", "Contracted Hours - Start Time",
    "Contracted Hours - Finish Time", "Contracted Hours - Role",
    "Contracted Hours - Department/Room", "Contracted Hours - Number Hours",
    "Contracted Hours - Number Days", "Contracted Hours - Cycle Start Date",
]

EMPLOYEE_CODE_COL = "Employee Code"
BALANCE_OWING_COL = "Leave - Balance Owing"
LEAVE_TYPE_COL    = "Leave - Type Code"


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def parse_smartly_csv(filepath: str) -> list[dict]:
    """Read the Smartly CSV, skipping the metadata header rows."""
    with open(filepath, encoding="utf-8-sig", newline="") as f:
        raw = f.read()

    lines = raw.splitlines(keepends=True)
    data_section = "".join(lines[SMARTLY_HEADER_SKIP:])
    reader = csv.DictReader(io.StringIO(data_section))
    return list(reader)


def to_float(value: str) -> float | None:
    """Convert a CSV string to float, returning None if blank or invalid."""
    if value is None:
        return None
    v = value.strip()
    if not v:
        return None
    try:
        return float(v)
    except ValueError:
        return None


def transform(rows: list[dict]) -> list[dict]:
    """
    Filter and reshape Smartly rows into Simplifi import rows.
    Returns a list of dicts keyed by Simplifi column name.
    """
    output = []

    for row in rows:
        leave_type = (row.get("Leave Type") or "").strip()

        if leave_type not in LEAVE_TYPE_MAP:
            continue  # skip unrequired leave types

        simplifi_code, balance_col = LEAVE_TYPE_MAP[leave_type]
        balance_value = to_float(row.get(balance_col))

        if balance_value is None:
            # Skip rows where the relevant balance is missing
            continue

        employee_code = (row.get("Employee Code") or "").strip()
        if not employee_code:
            continue

        out_row = {col: None for col in SIMPLIFI_COLUMNS}
        out_row[EMPLOYEE_CODE_COL] = employee_code
        out_row[BALANCE_OWING_COL] = balance_value
        out_row[LEAVE_TYPE_COL]    = simplifi_code
        output.append(out_row)

    return output


def write_xlsx(output_rows: list[dict], output_path: str) -> None:
    """Write the transformed rows to a Simplifi-formatted Excel file."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Staff Import Template"

    # Header row
    ws.append(SIMPLIFI_COLUMNS)

    # Data rows
    for row in output_rows:
        ws.append([row.get(col) for col in SIMPLIFI_COLUMNS])

    # Basic formatting — bold the header row
    from openpyxl.styles import Font
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Auto-width the three populated columns for readability
    for col_name in [EMPLOYEE_CODE_COL, BALANCE_OWING_COL, LEAVE_TYPE_COL]:
        col_idx = SIMPLIFI_COLUMNS.index(col_name) + 1
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(col_idx)
        ].width = 22

    wb.save(output_path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    if len(sys.argv) < 2:
        print(__doc__)
        print("ERROR: Please provide an input CSV path as the first argument.")
        sys.exit(1)

    input_csv = sys.argv[1]

    if not os.path.isfile(input_csv):
        print(f"ERROR: File not found: {input_csv}")
        sys.exit(1)

    # Determine output path
    if len(sys.argv) >= 3:
        output_xlsx = sys.argv[2]
    else:
        stem = Path(input_csv).stem
        output_xlsx = str(Path(input_csv).parent / f"{stem}_simplifi_import.xlsx")

    print(f"Reading:  {input_csv}")
    rows = parse_smartly_csv(input_csv)
    print(f"  → {len(rows)} source rows loaded")

    output_rows = transform(rows)
    print(f"  → {len(output_rows)} leave balance rows after filtering")

    write_xlsx(output_rows, output_xlsx)
    print(f"Written:  {output_xlsx}")

    # Summary by leave type
    from collections import Counter
    counts = Counter(r[LEAVE_TYPE_COL] for r in output_rows)
    print("\nSummary:")
    for lt, count in sorted(counts.items()):
        print(f"  {lt}: {count} employee records")


if __name__ == "__main__":
    main()
